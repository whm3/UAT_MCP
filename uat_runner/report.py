"""Generate UAT reports in markdown and XLSX formats."""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496",
                          fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="1F3864", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                           fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                        fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                        fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                        fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def generate_markdown_report(session, output_path):
    """Generate a markdown report from a completed session.

    Args:
        session: Session object with results.
        output_path: Path to write the markdown file.
    """
    started = datetime.fromisoformat(session.started)
    updated = datetime.fromisoformat(session.updated)
    duration = updated - started
    hours, remainder = divmod(int(duration.total_seconds()), 3600)
    minutes = remainder // 60

    if hours > 0:
        duration_str = f"{hours}h {minutes}m"
    else:
        duration_str = f"{minutes}m"

    overall = "PASS" if session.failed == 0 else "FAIL"

    lines = [
        f"# UAT Report: {session.plan_name} v{session.version}",
        "",
        "| Field | Value |",
        "|-------|-------|",
        f"| Plan | {session.plan_name} |",
        f"| Version | {session.version} |",
        f"| Tester | {session.tester} |",
        f"| Date | {started.strftime('%Y-%m-%d')} |",
        f"| Duration | {duration_str} |",
        f"| Result | **{overall}**"
        + (f" ({session.failed} failure{'s' if session.failed != 1 else ''})"
           if session.failed > 0 else "") + " |",
        "",
        "## Summary",
        "",
    ]

    total = session.total
    if total > 0:
        lines.append(
            f"- **Pass:** {session.passed} ({session.passed/total*100:.1f}%)")
        lines.append(
            f"- **Fail:** {session.failed} ({session.failed/total*100:.1f}%)")
        lines.append(
            f"- **Skip:** {session.skipped} ({session.skipped/total*100:.1f}%)")
    else:
        lines.append("No tests found.")
    lines.append("")

    # Failures section
    if session.failed > 0:
        lines.append("## Failures")
        lines.append("")
        lines.append("| # | Section | Test | Comment |")
        lines.append("|---|---------|------|---------|")
        for section in session.sections:
            for test in section.tests:
                if test.result == "fail":
                    comment = test.comment.replace("|", "\\|") if test.comment else ""
                    test_text = test.test.replace("|", "\\|")
                    lines.append(
                        f"| {test.num} | {section.title} | "
                        f"{test_text} | {comment} |")
        lines.append("")

    # Skipped section
    if session.skipped > 0:
        lines.append("## Skipped")
        lines.append("")
        lines.append("| # | Section | Test | Comment |")
        lines.append("|---|---------|------|---------|")
        for section in session.sections:
            for test in section.tests:
                if test.result == "skip":
                    comment = test.comment.replace("|", "\\|") if test.comment else ""
                    test_text = test.test.replace("|", "\\|")
                    lines.append(
                        f"| {test.num} | {section.title} | "
                        f"{test_text} | {comment} |")
        lines.append("")

    # Full results by section
    lines.append("## Full Results")
    lines.append("")
    for section in session.sections:
        lines.append(f"### {section.title}")
        lines.append("")
        lines.append("| # | Test | Expected | Result | Comment |")
        lines.append("|---|------|----------|--------|---------|")
        for test in section.tests:
            result = (test.result or "—").upper()
            comment = test.comment.replace("|", "\\|") if test.comment else ""
            test_text = test.test.replace("|", "\\|")
            expected = test.expected.replace("|", "\\|")
            lines.append(
                f"| {test.num} | {test_text} | {expected} | "
                f"{result} | {comment} |")
        lines.append("")

    with open(output_path, "w") as f:
        f.write("\n".join(lines))


def generate_xlsx_report(session, output_path):
    """Generate an XLSX report from a completed session.

    Args:
        session: Session object with results.
        output_path: Path to write the XLSX file.
    """
    wb = Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 40

    summary_rows = [
        ("Plan", session.plan_name),
        ("Version", session.version),
        ("Tester", session.tester),
        ("Started", session.started),
        ("Completed", session.completed or "In progress"),
        ("", ""),
        ("Total Tests", session.total),
        ("Pass", session.passed),
        ("Fail", session.failed),
        ("Skip", session.skipped),
        ("Remaining", session.remaining),
    ]

    for row_idx, (label, value) in enumerate(summary_rows, 1):
        cell_a = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_b = ws_sum.cell(row=row_idx, column=2, value=value)
        cell_a.font = Font(bold=True)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER

    # Results sheet
    ws = wb.create_sheet(title="Results")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 35
    ws.freeze_panes = "A2"

    headers = ["#", "Test", "Expected", "Result", "Comment"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    row = 2
    for section in session.sections:
        # Section header row
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=section.title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        # Test rows
        for test in section.tests:
            ws.cell(row=row, column=1, value=test.num).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="center", vertical="top")

            test_cell = ws.cell(row=row, column=2, value=test.test)
            test_cell.alignment = WRAP
            test_cell.border = THIN_BORDER

            exp_cell = ws.cell(row=row, column=3, value=test.expected)
            exp_cell.alignment = WRAP
            exp_cell.border = THIN_BORDER

            result_val = (test.result or "").upper()
            result_cell = ws.cell(row=row, column=4, value=result_val)
            result_cell.border = THIN_BORDER
            result_cell.alignment = Alignment(
                horizontal="center", vertical="top")

            if test.result == "pass":
                result_cell.fill = PASS_FILL
            elif test.result == "fail":
                result_cell.fill = FAIL_FILL
            elif test.result == "skip":
                result_cell.fill = SKIP_FILL

            comment_cell = ws.cell(row=row, column=5,
                                   value=test.comment or "")
            comment_cell.border = THIN_BORDER
            comment_cell.alignment = WRAP

            row += 1

        # Blank separator
        row += 1

    # Failures sheet (if any)
    if session.failed > 0:
        ws_fail = wb.create_sheet(title="Failures")
        ws_fail.column_dimensions["A"].width = 10
        ws_fail.column_dimensions["B"].width = 30
        ws_fail.column_dimensions["C"].width = 55
        ws_fail.column_dimensions["D"].width = 55
        ws_fail.column_dimensions["E"].width = 35
        ws_fail.freeze_panes = "A2"

        fail_headers = ["#", "Section", "Test", "Expected", "Comment"]
        for col, h in enumerate(fail_headers, 1):
            cell = ws_fail.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="C00000", end_color="C00000",
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        fail_row = 2
        for section in session.sections:
            for test in section.tests:
                if test.result == "fail":
                    ws_fail.cell(row=fail_row, column=1,
                                 value=test.num).border = THIN_BORDER
                    ws_fail.cell(row=fail_row, column=2,
                                 value=section.title).border = THIN_BORDER
                    ws_fail.cell(row=fail_row, column=3,
                                 value=test.test).border = THIN_BORDER
                    ws_fail.cell(row=fail_row, column=4,
                                 value=test.expected).border = THIN_BORDER
                    ws_fail.cell(row=fail_row, column=5,
                                 value=test.comment or "").border = THIN_BORDER
                    for c in range(1, 6):
                        ws_fail.cell(row=fail_row, column=c).alignment = WRAP
                        ws_fail.cell(row=fail_row, column=c).fill = FAIL_FILL
                    fail_row += 1

    wb.save(output_path)
